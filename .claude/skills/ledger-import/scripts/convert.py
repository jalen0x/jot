"""Convert Alipay CSV / WeChat XLSX / CMB PDF to the app's JSON import schema.

Run from the project root via uv (handles openpyxl + pdftotext shell-out):
  uv run --with openpyxl --no-project python3 \\
    .claude/skills/ledger-import/scripts/convert.py            # all three
  uv run --with openpyxl --no-project python3 \\
    .claude/skills/ledger-import/scripts/convert.py alipay     # just one source

Output goes to tmp/import/<source>.json (jot.jalenx.me ImportBatch schema).
Customize ACCOUNT_NAME, ALIPAY_CATEGORY_MAP, KEYWORD_CATEGORY,
categorize_personal_transfer, and CMB_DROP_SUMMARIES below.
See .claude/skills/ledger-import/SKILL.md and REFERENCE.md for the rules.

Comment policy: just the counterparty (or merchant for 淘宝闪购) — no source
prefix, no payment instrument, no transaction-type tag, no legal-entity suffix,
no account numbers from 招行 PDF.
"""
from __future__ import annotations
import csv, io, json, re, subprocess, sys
from datetime import datetime
from pathlib import Path

ACCOUNT_NAME = "内地账户"
TZ_MINUTES = 480  # UTC+08:00
TZ_SUFFIX = "+08:00"
DEFAULT_EXPENSE_CATEGORY = "其他"
DEFAULT_INCOME_CATEGORY = "其它"

DESKTOP = Path.home() / "Desktop"
OUT_DIR = Path("tmp/import")

# ---------- coarse category mapping ----------

ALIPAY_CATEGORY_MAP = {
    "餐饮美食": "餐饮",
    "服饰装扮": "服饰",
    "充值缴费": "话费",
    "日用百货": "日用",
    "交通出行": "交通",
    "数码电器": "数码",
    "美容美发": "美容",
    "商业服务": "虚拟服务",
    "学习教育": "学习",
    "医疗健康": "医疗",
    "文娱休闲": "娱乐",
    "运动户外": "其他",
    "居家家装": "居家",
    "鲜花宠物": "礼物",
    "投资理财": "其他",
    "其他": "其他",
}

KEYWORD_CATEGORY = [
    (re.compile(r"蜜雪|喜茶|coco|奈雪|霸王茶姬|益禾堂|沪上阿姨", re.I), "奶茶"),
    (re.compile(r"麦当劳|肯德基|kfc|星巴克|海底捞|美团|饿了么|拉扎斯|盒马|食堂|餐厅|餐饮", re.I), "餐饮"),
    (re.compile(r"滴滴|哈啰|青桔|铁路12306|高铁|地铁|出租车|加油"), "交通"),
    (re.compile(r"中国移动|中国电信|中国联通|话费充值|话费"), "话费"),
    (re.compile(r"京东|淘宝|拼多多|天猫"), "购物"),
    (re.compile(r"礼品卡|兑换码|订阅|会员|开通"), "虚拟服务"),
    (re.compile(r"红包|发给"), "亲友"),
    (re.compile(r"工资|代发"), "工资"),
]

# ---------- comment cleanup ----------

COMPANY_SUFFIXES = [
    re.compile(r"集团有限公司$"),
    re.compile(r"有限公司$"),
    re.compile(r"\s*[（(]个体工商户[）)]\s*$"),
]
PLATFORM_AGGREGATOR = re.compile(r"^淘宝闪购$")
ORDER_SUFFIX = re.compile(r"(?:外卖订单|订单)$")
CMB_ACCOUNT_NUMBER = re.compile(r"\d{8,}")
CMB_TRAILING_PAREN = re.compile(r"\s*[（(][^）)]*[）)]\s*$")
CMB_HEADER_NOISE = re.compile(
    r"招商银行|交易流水|Transaction|Statement|户\s*名|账号|Name|Account|账户类型|开\s*户\s*行|"
    r"Sub Branch|申请时间|Date|验\s*证\s*码|Verification|记账日期|货币|交易金额|联机余额|"
    r"交易摘要|对手信息|Currency|Amount|Balance|Counter Party|Type|页码|Page|"
    r"China Merchants"
)

# CMB summary types whose rows duplicate Alipay/WeChat bank-side entries. Drop them.
CMB_DROP_SUMMARIES = {"快捷支付", "银联快捷支付", "网联收款", "银联代付"}


def strip_company_suffix(s: str) -> str:
    s = s.strip()
    for pat in COMPANY_SUFFIXES:
        s = pat.sub("", s)
    return s.strip()


def format_party_comment(party: str, product: str) -> str:
    """Build the cleaned comment from 对方 + 商品 (Alipay/WeChat schema)."""
    party = (party or "").strip()
    product = (product or "").strip()
    if product == "/":
        product = ""
    if PLATFORM_AGGREGATOR.match(party) and product:
        out = ORDER_SUFFIX.sub("", product).strip()
    else:
        out = party or product
    return strip_company_suffix(out)


def format_cmb_comment(summary: str, counter_blob: str) -> str:
    """Build the cleaned comment for a 招行 row (摘要 + counter, no account numbers)."""
    blob = f"{summary} {counter_blob}".strip()
    blob = CMB_ACCOUNT_NUMBER.sub(" ", blob)
    blob = re.sub(r"\s+", " ", blob).strip()
    while CMB_TRAILING_PAREN.search(blob):
        blob = CMB_TRAILING_PAREN.sub("", blob).strip()
    return strip_company_suffix(blob)


# ---------- category resolution ----------


def categorize_personal_transfer(income: bool, party: str) -> str:
    """Personal red-packet / transfer flows route to 莹莹-specific categories when
    the counterparty is 莹莹, otherwise to the income/expense defaults so the
    category-type validator stays happy."""
    if "莹莹" in (party or ""):
        return "❤️莹莹" if income else "莹莹❤️"
    return DEFAULT_INCOME_CATEGORY if income else "亲友"


def coarse_category(income: bool, alipay_cat: str | None, text: str) -> str:
    if alipay_cat:
        mapped = ALIPAY_CATEGORY_MAP.get(alipay_cat.strip())
        if mapped:
            return mapped
    for regex, cat in KEYWORD_CATEGORY:
        if regex.search(text):
            return cat
    return DEFAULT_INCOME_CATEGORY if income else DEFAULT_EXPENSE_CATEGORY


# ---------- shared helpers ----------


def amount_to_cents(amount_str: str) -> int:
    return int(round(float(amount_str.strip().replace(",", "")) * 100))


def iso8601(ts: str) -> str:
    dt = datetime.strptime(ts.strip(), "%Y-%m-%d %H:%M:%S")
    return dt.strftime("%Y-%m-%dT%H:%M:%S") + TZ_SUFFIX


def make_row(*, ts: str, kind: str, category: str, cents: int, comment: str) -> dict:
    return {
        "transacted_at": iso8601(ts),
        "timezone_utc_offset_minutes": TZ_MINUTES,
        "transaction_kind": kind,
        "account_name": ACCOUNT_NAME,
        "transaction_category_name": category,
        "source_amount_cents": cents,
        "destination_amount_cents": 0,
        "comment": comment,
        "transaction_tag_names": [],
    }


# ---------- Alipay ----------


def convert_alipay() -> tuple[list[dict], list[str]]:
    src = next(DESKTOP.glob("支付宝交易明细*.csv"))
    raw = src.read_bytes().decode("gbk", errors="replace")
    lines = raw.splitlines()
    header_idx = next(i for i, line in enumerate(lines) if line.startswith("交易时间,"))
    body = "\n".join(lines[header_idx:])
    reader = csv.DictReader(io.StringIO(body))
    rows: list[dict] = []
    notes: list[str] = []
    for r in reader:
        flow = (r.get("收/支") or "").strip()
        if flow not in ("支出", "收入"):
            notes.append(f"SKIP (alipay 不计收支) {r.get('交易时间','').strip()} {r.get('商品说明','').strip()} {r.get('金额','').strip()}")
            continue
        income = flow == "收入"
        party = (r.get("交易对方") or "").strip()
        product = (r.get("商品说明") or "").strip()
        text = " ".join(filter(None, [party, product]))

        if "莹莹" in party:
            cat = categorize_personal_transfer(income, party)
        else:
            cat = coarse_category(income, r.get("交易分类"), text)

        rows.append(make_row(
            ts=r["交易时间"],
            kind="income" if income else "expense",
            category=cat,
            cents=amount_to_cents(r["金额"]),
            comment=format_party_comment(party, product),
        ))
    return rows, notes


# ---------- WeChat ----------


def convert_wechat() -> tuple[list[dict], list[str]]:
    import openpyxl
    src = next(DESKTOP.glob("微信支付账单流水文件*.xlsx"))
    wb = openpyxl.load_workbook(src, data_only=True)
    ws = wb.active
    header_row = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if row and row[0] == "交易时间":
            header_row = i
            break
    if header_row is None:
        raise SystemExit("could not find header in WeChat xlsx")
    header = [c.strip() if isinstance(c, str) else c for c in next(ws.iter_rows(min_row=header_row + 1, max_row=header_row + 1, values_only=True))]

    rows: list[dict] = []
    notes: list[str] = []
    for row in ws.iter_rows(min_row=header_row + 2, values_only=True):
        if not row or row[0] is None:
            continue
        rec = dict(zip(header, row))
        flow = (rec.get("收/支") or "").strip()
        if flow not in ("支出", "收入"):
            notes.append(f"SKIP (wechat 中性交易) {rec.get('交易时间')} {rec.get('交易对方')} {rec.get('金额(元)')}")
            continue
        income = flow == "收入"
        party = str(rec.get("交易对方") or "").strip()
        product = str(rec.get("商品") or "").strip()
        wechat_type = (rec.get("交易类型") or "").strip()
        text = " ".join(filter(None, [party, product, wechat_type]))

        if "莹莹" in party or "红包" in wechat_type or "转账" in wechat_type:
            cat = categorize_personal_transfer(income, party)
        else:
            cat = coarse_category(income, None, text)

        ts = rec["交易时间"]
        ts_str = ts.strftime("%Y-%m-%d %H:%M:%S") if isinstance(ts, datetime) else str(ts).strip()
        amount = rec.get("金额(元)") or rec.get("金额")
        amount_str = str(amount).strip().lstrip("¥").lstrip("￥")

        rows.append(make_row(
            ts=ts_str,
            kind="income" if income else "expense",
            category=cat,
            cents=amount_to_cents(amount_str),
            comment=format_party_comment(party, product),
        ))
    return rows, notes


# ---------- CMB ----------


CMB_ROW = re.compile(r"^(\d{4}-\d{2}-\d{2})\s+(CNY|USD|HKD)\s+(-?[\d,]+\.\d{2})\s+([\d,]+\.\d{2})\s+(\S+)\s*(.*)$")


def convert_cmb() -> tuple[list[dict], list[str]]:
    src = next(DESKTOP.glob("招商银行交易流水*.pdf"))
    text = subprocess.check_output(["pdftotext", "-layout", str(src), "-"]).decode("utf-8")

    items: list[tuple[str, object]] = []
    for raw_line in text.splitlines():
        line = raw_line.strip()
        if not line or CMB_HEADER_NOISE.search(line):
            continue
        m = CMB_ROW.match(line)
        if m:
            items.append(("DATA", m.groups()))
        elif re.search(r"[一-鿿]|^\d{6,}$", line):
            items.append(("CTX", line))

    rows: list[dict] = []
    notes: list[str] = []
    pending: list[str] = []
    for kind, payload in items:
        if kind == "CTX":
            pending.append(payload)
            continue
        date, currency, amount, balance, summary, counter_inline = payload
        pre_counter = pending[:]
        pending = []

        if summary in CMB_DROP_SUMMARIES:
            notes.append(f"SKIP (cmb dedup w/ alipay/wechat) {date} {summary} {(counter_inline or '').strip()}")
            continue

        cents_signed = amount_to_cents(amount)
        tx_kind = "income" if cents_signed >= 0 else "expense"
        income = tx_kind == "income"
        cents = abs(cents_signed)
        if cents == 0:
            notes.append(f"SKIP (cmb zero amount) {date} {summary}")
            continue

        counter_blob = " ".join(filter(None, pre_counter + [counter_inline.strip()]))
        text_blob = f"{summary} {counter_blob}"

        if income and re.search(r"工资|代发", text_blob):
            cat = "工资"
        elif "ATM" in summary or "取款" in summary:
            cat = "其他"
        elif "转账" in summary or "代付" in summary or "行内转账" in summary:
            cat = categorize_personal_transfer(income, counter_blob)
        else:
            cat = coarse_category(income, None, text_blob)

        rows.append(make_row(
            ts=f"{date} 00:00:00",
            kind=tx_kind,
            category=cat,
            cents=cents,
            comment=format_cmb_comment(summary, counter_blob),
        ))

    return rows, notes


# ---------- driver ----------


def main(argv: list[str]) -> int:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    cmd = argv[1] if len(argv) > 1 else "all"
    handlers = {
        "alipay": convert_alipay,
        "wechat": convert_wechat,
        "cmb": convert_cmb,
    }
    sources = [cmd] if cmd in handlers else list(handlers)
    for name in sources:
        rows, notes = handlers[name]()
        out_path = OUT_DIR / f"{name}.json"
        out_path.write_text(json.dumps({"transactions": rows}, ensure_ascii=False, indent=2))
        notes_path = OUT_DIR / f"{name}.skipped.txt"
        notes_path.write_text("\n".join(notes) + ("\n" if notes else ""))
        print(f"{name}: {len(rows)} rows -> {out_path}; {len(notes)} skipped -> {notes_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))
